"""loads scrapper data to db"""

import argparse
import pymysql
import openpyxl
import xlrd
import os
from pathlib import Path
from datetime import datetime

from loguru import logger

from court_cases_scrapper import config


def load_to_dm():
    """calls load to dm procedure"""

    logger.debug("Connecting to db")

    conn = pymysql.connect(host=config.MYSQL_CONNECT["host"],
                           port=config.MYSQL_CONNECT["port"],
                           user=config.MYSQL_CONNECT["user"],
                           passwd=config.MYSQL_CONNECT["passwd"]
                           )

    logger.debug("Connected")
    cursor = conn.cursor()

    sql = "call dm.p_load_court_cases()"
    cursor.execute(sql)
    conn.commit()
    logger.info("DM data updated")
    cursor.close()

def check_filter_condition(file_name: str, date_from: str, date_to: str, court_filter: str) -> bool:
    """checks if condition met"""
    if not date_from and not date_to and not court_filter:
        return True
    parts = file_name.split("_")
    court_name = parts[0]
    file_date = datetime.strptime(parts[1], "%Y%m")

    if date_from:
        d_from = datetime.strptime(date_from, "%Y-%m")
    else:
        d_from = datetime(year=2020, month=1, day=1)

    if date_to:
        d_to = datetime.strptime(date_to, "%Y-%m")
    else:
        d_to = datetime(year=2030, month=12, day=1)

    if (not court_filter or court_filter == court_name) and (d_from <= file_date <= d_to):
        return True
    else:
        return False


def load_stage(date_from: str, date_to: str, court_filter: str, output_mode: str) -> None:
    """reads file list from dir and calls load method"""

    logger.debug("Connecting to db")

    conn = pymysql.connect(host=config.MYSQL_CONNECT["host"],
                           port=config.MYSQL_CONNECT["port"],
                           user=config.MYSQL_CONNECT["user"],
                           passwd=config.MYSQL_CONNECT["passwd"]
                           )

    logger.debug("Connected")
    cursor = conn.cursor()

    logger.debug("Cleaning stage table " + config.STAGE_TABLE)
    sql = "delete from " + config.STAGE_TABLE
    cursor.execute(sql)
    logger.debug("Stage table cleaned")
    sql_statement = f"INSERT INTO {config.STAGE_TABLE} ("

    for field in config.STAGE_STRUCTURE:
        sql_statement = sql_statement + field + ", "

    sql_statement = sql_statement.rstrip(", ") + ") VALUES ("

    for field in config.STAGE_STRUCTURE:
        if field == "load_dttm":
            sql_statement = sql_statement + "now(), "
        else:
            sql_statement = sql_statement + "%s, "

    sql_statement = sql_statement.rstrip(", ") + ")"

    logger.debug(sql_statement)
    cursor.close()
    logger.info("Data load start")
    for file_str in os.listdir(config.OUTPUT_DIR):
        file = os.path.join(config.OUTPUT_DIR, file_str)
        if os.path.isfile(Path(file)) and file_str.endswith(output_mode):
            if check_filter_condition(file_str, date_from, date_to, court_filter):
                if output_mode == "xlsx":
                    load_xlsx_file_to_stage(Path(file), conn, sql_statement)
                elif output_mode == "xls":
                    load_xls_file_to_stage(Path(file), conn, sql_statement)

    logger.info("Data load completed")
    conn.close()


def load_xls_file_to_stage(file_name: Path, conn: pymysql.connect, sql_statement: str) -> None:
    """load xls file to stage"""
    logger.debug("Loading " + str(file_name))
    workbook = xlrd.open_workbook_xls(file_name)

    sheet = workbook.sheet_by_index(0)

    cursor = conn.cursor()

    for row in range(1, sheet.nrows):
        values = []
        for col in range(0, sheet.ncols):
            if sheet.cell(row, col).value == "":
                values.append(None)
            else:
                values.append(str(sheet.cell(row, col).value).strip())

        cursor.execute(sql_statement, values)
        if row % config.COMMIT_INTERVAL == 0:
            conn.commit()
            logger.debug("Commit " + str(row))

    logger.debug("File " + str(file_name) + " loaded")
    conn.commit()
    cursor.close()


def load_xlsx_file_to_stage(file_name: Path, conn: pymysql.connect, sql_statement: str) -> None:
    """load xls file to stage"""
    logger.debug("Loading " + str(file_name))
    workbook = openpyxl.load_workbook(file_name)

    sheet = workbook.worksheets[0]

    cursor = conn.cursor()

    for row in range(1, sheet.max_row):
        values = []
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[row].value == "":
                values.append(None)
            else:
                values.append(str(col[row].value).strip())

        cursor.execute(sql_statement, values)
        if row % config.COMMIT_INTERVAL == 0:
            conn.commit()
            logger.debug("Commit " + str(row))

    workbook.close()

    logger.debug("File " + str(file_name) + " loaded")
    conn.commit()
    cursor.close()


def parse_args() -> argparse.Namespace:
    """Parser for command-line options, arguments and sub-commands."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--court", type=str, required=False)
    parser.add_argument("--output_mode", type=str, required=False, default="xls", choices=("xls", "xlsx"))
    parser.add_argument("--date_from", type=str, required=False)
    parser.add_argument("--date_to", type=str, required=False)
    parsed_args = parser.parse_args()

    return parsed_args


def main() -> None:
    args = parse_args()
    load_stage(args.date_from, args.date_to, args.court, args.output_mode)
    load_to_dm()


if __name__ == "__main__":
    main()
