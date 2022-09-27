"""
parse court data
"""
import argparse
import openpyxl
import xlwt
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup
from loguru import logger

from court_cases_scrapper import config


def save_to_excel(result: list[dict[str, str]], period: str, file_name_part: str, output_mode: str) -> None:
    if output_mode == "xlsx":
        save_to_xlsx(result, period, file_name_part)
    elif output_mode == "xls":
        save_to_xls(result, period, file_name_part)


def save_to_xls(data: list[dict[str, str]], period: str, file_name_part: str) -> None:
    """creates xls file"""
    workbook = xlwt.Workbook()

    sheet = workbook.add_sheet("Court Cases")
    for idx, column in enumerate(config.EXCEL_TABLE_CONFIG):
        sheet.write(0, idx, column.get("title"))

    for idx_r, row in enumerate(data):
        for idx_c, col in enumerate(config.EXCEL_TABLE_CONFIG):
            sheet.write(idx_r+1, idx_c, row.get(col.get("content")))

    current_dttm = datetime.now().strftime("%Y%m%d%H%M%S")
    workbook.save(f"{config.OUTPUT_DIR}/{file_name_part}_{period}__{current_dttm}.xls")


def create_xls_row(data: dict[str, str], mode: str) -> list[str]:
    """forms string for xlsx file"""
    row = []
    if mode == "header":
        for column in config.EXCEL_TABLE_CONFIG:
            row.append(column.get("title"))
    else:
        for column in config.EXCEL_TABLE_CONFIG:
            row.append(data.get(column.get("content")))

    return row


def save_to_xlsx(data: list[dict[str, str]], period: str, file_name_part: str) -> None:
    """save to xlsx"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Court cases " + period

    header = create_xls_row({}, "header")
    sheet.append(header)
    for row in data:
        row = create_xls_row(row, "data")
        sheet.append(row)

    # resize columns
    for column in config.EXCEL_TABLE_CONFIG:
        sheet.column_dimensions[column.get("position")].width = column.get("width")

    # hyperlinks
    for column in config.EXCEL_TABLE_CONFIG:
        if column.get("style") == "Hyperlink":
            for cell in sheet[column.get("position")]:
                if cell.row > 1 and cell.value:
                    cell.style = "Hyperlink"
                    cell.hyperlink = str(cell.value)

    # add style
    table = Table(
        displayName="Court_cases", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
    )
    style = TableStyleInfo(
        name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    # save file
    current_dttm = datetime.now().strftime("%Y%m%d%H%M%S")
    workbook.save(f"{config.OUTPUT_DIR}/{file_name_part}_{period}__{current_dttm}.xlsx")


def daterange(date1: str, date2: str) -> list[datetime]:
    d_date1 = datetime.strptime(date1, '%Y-%m-%d')
    d_date2 = datetime.strptime(date2, '%Y-%m-%d')
    result = []
    for n in range(int((d_date2 - d_date1).days) + 1):
        result.append((d_date1 + timedelta(days=n)))

    return result


def parse_page(page: requests.Response, court: dict, check_date: str) -> list[dict[str, str]]:
    """parses output page"""
    result = []
    soup = BeautifulSoup(page.content, 'html.parser')
    tables = soup.find_all("div", id="tablcont")
    logger.info("Processing " + court.get("file_name_part") + ", date " + check_date)
    # <div id="tablcont">
    for table in tables:
        section_name = ""
        sections = table.find_all("tr")
        # tr
        for idx, section in enumerate(sections):
            if idx == 0:
                continue
            # setting new section
            if len(section.contents) == 1:
                for idx_r, row in enumerate(section.find_all("td")):
                    section_name = row.text
            # appending row
            else:
                result_row = {"section_name": section_name}
                # td
                for idx_r, row in enumerate(section.find_all("td")):
                    if row.text:
                        result_row["col" + str(idx_r)] = row.text
                    else:
                        result_row["col" + str(idx_r)] = str(row.contents)
                    if row.find(href=True):
                        result_row["col" + str(idx_r) + "_link"] = court.get("link") + row.find(href=True)["href"]

                result_row["check_date"] = check_date
                result_row["court"] = court.get("title")
                result.append(result_row)
    return result


def scrap_spb_district_courts(date_from: str, date_to: str, court_filter: str, output_mode: str):
    session = requests.Session()
    session.headers = {"user-agent": config.USER_AGENT}
    for court in config.SPD_DISTRICT_COURTS:
        period = datetime.strptime(date_from, '%Y-%m-%d').strftime("%Y%m")
        if (court_filter and court.get("file_name_part") != court_filter) or (court.get("skip")):
            continue
        result = []
        for date in daterange(date_from, date_to):
            # save file if new month
            if period != date.strftime("%Y%m"):
                save_to_excel(result, period, court.get("file_name_part"), output_mode)
                period = date.strftime("%Y%m")
                result = []

            check_date = date.strftime("%d.%m.%Y")
            page = session.get(court.get("link") + "/modules.php?name=sud_delo&srv_num=1&H_date=" + check_date)
            result = result + parse_page(page, court, check_date)
        save_to_excel(result, period, court.get("file_name_part"), output_mode)


def parse_args() -> argparse.Namespace:
    """Parser for command-line options, arguments and sub-commands."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--court", type=str, required=False)
    parser.add_argument("--output_mode", type=str, required=False, default="xls", choices=("xls", "xlsx"))
    parser.add_argument("--date_from", type=str, required=False, default=datetime.now().strftime("%Y-%m-%d"))
    parser.add_argument("--date_to", type=str, required=False,
                        default=(datetime.now() + timedelta(days=60)).strftime("%Y-%m-%d"))
    parsed_args = parser.parse_args()

    return parsed_args


def main() -> None:
    args = parse_args()
    scrap_spb_district_courts(args.date_from, args.date_to, args.court, args.output_mode)


if __name__ == "__main__":
    main()
